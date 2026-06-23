import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import networkx as nx
import joblib
import io

st.set_page_config(page_title="Прогноз академических рисков", layout="wide")

st.title("Прогнозирование академических рисков студентов")
st.markdown("Загрузите ведомость успеваемости в формате Excel и выберите семестр для оценки вероятности отчисления.")

uploaded_file = st.file_uploader("Шаг 1: загрузите Excel-файл успеваемости", type=["xlsx"])
semester_choice = st.selectbox("Шаг 2: выберите семестр для прогнозирования", [1, 2, 3, 4, 5, 6])

checkpoints = {1: 7, 2: 15, 3: 22, 4: 29, 5: 38, 6: 47}
col_limit = checkpoints[semester_choice]

@st.cache_data
def load_template():
    df_temp = pd.read_excel('data/ПМИ4_общий_датасет.xlsx', index_col=0)
    df_temp.columns = [str(col).strip() for col in df_temp.columns]
    return df_temp

@st.cache_resource
def load_graph():
    graph_df = pd.read_excel('data/Граф знаний.xlsx', header=None, names=['Source', 'Target'])
    graph_df.dropna(inplace=True)
    graph_df['Source'] = graph_df['Source'].astype(str).str.strip()
    graph_df['Target'] = graph_df['Target'].astype(str).str.strip()
    
    G = nx.DiGraph()
    G.add_edges_from(zip(graph_df['Source'], graph_df['Target']))
    return G

@st.cache_resource
def load_ml_model(semester):
    model_path = f"models/model_semester_{semester}.pkl"
    try:
        return joblib.load(model_path)
    except FileNotFoundError:
        st.error(f"Файл модели '{model_path}' не найден в папке проекта.")
        return None

def capitalize_first(s):
    if s:
        return s[0].upper() + s[1:]
    return s

def interpret_risk(student_grades, G):
    interpretation = {
        'Приоритетные предметы': []
    }

    if isinstance(student_grades, np.ndarray):
        student_grades = pd.Series(student_grades)

    failing_subjects = {}
    for subj, grade in student_grades.items():
        if pd.notna(grade) and grade < 3:
            failing_subjects[subj] = grade
    
    triple_subjects = {}
    for subj, grade in student_grades.items():
        if pd.notna(grade) and grade == 3:
            triple_subjects[subj] = grade
    
    subject_priority = []
    
    for subj, grade in list(failing_subjects.items()) + list(triple_subjects.items()):
        if subj in G.nodes():
            n_dependents = len(list(G.successors(subj)))
            dependent_list = list(G.successors(subj))
            criticality = 2 if grade == 2 else 1
            priority = n_dependents * criticality
            if n_dependents >= 2:
                subject_priority.append((subj, grade, n_dependents, dependent_list, priority))

    subject_priority.sort(key=lambda x: x[4], reverse=True)
    
    if subject_priority:
        for subj, grade, n_dependents, dependent_list, _ in subject_priority:
            interpretation['Приоритетные предметы'].append({
                'предмет': subj,
                'оценка': grade,
                'количество_зависимых': n_dependents,
                'зависимые_дисциплины': dependent_list
            })
    
    return interpretation

if uploaded_file is not None:
    G = load_graph()
    model = load_ml_model(semester_choice)
    df_template = load_template()
    
    if model is not None and df_template is not None:
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        
        non_grades = ["", "-", "–", "—", "н/я", "н/з", "0", 0]
        subject_row = 3
        
        for row_idx in range(sheet.max_row, 3, -1):
            fio_cell = sheet.cell(row=row_idx, column=1).value
            if fio_cell is None or str(fio_cell).strip() == "":
                sheet.delete_rows(row_idx)
                continue
                
            row_grade_cells = [sheet.cell(row=row_idx, column=c) for c in range(2, sheet.max_column + 1)]
            is_row_empty = all(cell.value is None or str(cell.value).strip() in non_grades for cell in row_grade_cells)
            if is_row_empty:
                sheet.delete_rows(row_idx)

        for col_idx in range(sheet.max_column, 1, -1):
            subject_cell = sheet.cell(row=subject_row, column=col_idx)
            if subject_cell is None:
                sheet.delete_cols(col_idx)
                continue
                
            subject_value = str(subject_cell.value).strip() if subject_cell.value is not None else ""
            subject_lower = subject_value.lower()

            if any(phrase in subject_lower for phrase in ["физ.", "элективн"]):
                subject_cell.value = "Физическая культура и спорт"
                continue
            
            if subject_value == "":
                sheet.delete_cols(col_idx)
                continue
                
            if any(phrase in subject_lower for phrase in ["практика", "курсов", " кр"]):
                sheet.delete_cols(col_idx)
                continue
                
            real_grades_count = 0
            for row_idx in range(4, sheet.max_row + 1):
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    val_str = str(cell_val).strip()
                    if val_str not in non_grades:
                        real_grades_count += 1
                        
            if real_grades_count < 5:
                sheet.delete_cols(col_idx)

        purple_codes = ['FF674EA7', 'FF9900FF', 'FF00FF00']
        red_codes = ['FFFF0000', 'FFCC0000', 'FF980000']
        
        for row_idx in range(4, sheet.max_row + 1):
            for col_idx in range(2, sheet.max_column + 1):
                subject_cell = sheet.cell(row=subject_row, column=col_idx)
                if subject_cell is None:
                    continue
                subject_value = str(subject_cell.value).strip() if subject_cell.value is not None else ""
                if subject_value == "":
                    continue
                    
                cell = sheet.cell(row=row_idx, column=col_idx)
                is_empty = cell.value is None
                current_color = str(cell.fill.start_color.index) if cell.fill.start_color.index else ""
                
                if not is_empty:
                    val_str = str(cell.value).strip()
                    if val_str and val_str[0].isdigit():
                        cell.value = int(val_str[0])
                    if cell.value == 1 or str(cell.value).strip() == "1":
                        cell.value = 5
                
                if is_empty:
                    if current_color in purple_codes:
                        cell.value = 3
                    else:
                        cell.value = 2

                if current_color in red_codes:
                    cell.value = 2

        student_data = []
        for row_idx in range(4, sheet.max_row + 1):
            fio = sheet.cell(row=row_idx, column=1).value
            grades = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(2, sheet.max_column + 1)]
            student_data.append({"ФИО": fio, "Оценки": grades})
            
        columns = [str(sheet.cell(row=3, column=c).value).strip() for c in range(2, sheet.max_column + 1)]
        df_clean = pd.DataFrame([s["Оценки"] for s in student_data], columns=columns)
        df_clean.insert(0, "Студент", [s["ФИО"] for s in student_data])
        df_clean = df_clean.set_index("Студент")
        
        replacements = {
            "дифференциальное и интегральное исчисление": "Дифференциальное и интегральное исчисления",
            "диференциальные уравнения в частных производных": "Дифференциальные уравнения в частных производных",
            "доп главы элементарной математики": "Дополнительные главы элементарной математики",
            "история (история россии, всеобщая история)": "История",
            "математика 4 сем": "Математика",
            "математическая логика и теория алгоритмов (5 семестр)": "Математическая логика и теория алгоритмов",
            "пнид": "Проектная и научно-исследовательская деятельность",
            "русский язык и культура речи": "Русский язык",
            "проектирование информационных систем": "Программная инженерия",
            "сложность вычисленй": "Сложность вычислений",
            "теория информации и кодирование": "Теория информации и кодирования"
        }

        new_cols = []
        for col in df_clean.columns:
            col_clean = str(col).replace('\xa0', ' ').replace('  ', ' ').strip()
            col_lower = col_clean.lower()
            if col_lower in replacements:
                new_cols.append(replacements[col_lower])
            else:
                new_cols.append(capitalize_first(col_clean))
        df_clean.columns = new_cols

        expected_subjects = list(df_template.columns[:col_limit])
        uploaded_subjects = list(df_clean.columns)
        
        extra_subjects = [subj for subj in uploaded_subjects if subj not in expected_subjects]
        missing_subjects = [subj for subj in expected_subjects if subj not in uploaded_subjects]
        
        extra_subjects_to_show = [subj for subj in extra_subjects if subj not in ['None', 'nan', 'NaN', '']]
        
        if missing_subjects:
            st.error("Ошибка: В загруженном файле отсутствуют обязательные предметы учебного плана!")
            for subj in missing_subjects:
                st.write(f"- **{subj}**")
            st.info("Пожалуйста, добавьте эти предметы в ваш Excel-файл и загрузите файл повторно.")
            
        else:
            if extra_subjects:
                if extra_subjects_to_show:
                    st.warning(f"Из файла автоматически исключены предметы: {', '.join(extra_subjects_to_show)}")
                df_clean = df_clean.drop(columns=extra_subjects)
                
            X_base = df_clean[expected_subjects]
            current_subjects = list(X_base.columns)

            if semester_choice == 1:
                c_fail = []
                c_min = []
                for idx, row in X_base.iterrows():
                    fail_count = 0
                    grades_list = []
                    for subj in current_subjects:
                        if subj in G and G.out_degree(subj) >= 2:
                            grade = row[subj]
                            if pd.notna(grade):
                                grades_list.append(grade)
                                if grade < 3:
                                    fail_count += 1
                    c_fail.append(fail_count)
                    c_min.append(min(grades_list) if grades_list else 5)
                
                X_base['C_fail'] = c_fail
                X_base['C_min'] = c_min

            elif semester_choice in [5, 6]:
                descendants_count = {node: len(nx.descendants(G, node)) for node in G.nodes()}
                risk_grade_3_core = []
                risk_grade_2_core = []
                
                for idx, row in X_base.iterrows():
                    score_3 = 0
                    score_2 = 0
                    for subj in current_subjects:
                        if subj in G.nodes():
                            grade = row[subj]
                            if pd.notna(grade):
                                if grade == 3:
                                    score_3 += descendants_count[subj]
                                elif grade < 3:
                                    score_2 += descendants_count[subj]
                    risk_grade_3_core.append(score_3)
                    risk_grade_2_core.append(score_2)
                
                X_base['Risk_Grade_3_Core'] = risk_grade_3_core
                X_base['Risk_Grade_2_Core'] = risk_grade_2_core

            X_base.columns = X_base.columns.astype(str)
            
            risk_probabilities = model.predict_proba(X_base)[:, 0]
            
            results_df = pd.DataFrame(index=df_clean.index)
            results_df['Вероятность академического риска (%)'] = np.round(risk_probabilities * 100, 2)
            
            results_df = results_df[results_df['Вероятность академического риска (%)'] >= 50.0]
            
            st.subheader("Результаты анализа успеваемости")
            
            if not results_df.empty:
                results_df = results_df.sort_values(by='Вероятность академического риска (%)', ascending=False)
                
                st.dataframe(results_df, use_container_width=True)
                
                st.subheader("Детальный анализ рисков")
                
                for student_name, row in results_df.iterrows():
                    risk_prob = row['Вероятность академического риска (%)']
                    student_grades = df_clean.loc[student_name]
                    
                    interp = interpret_risk(student_grades, G)
                    
                    with st.expander(f"{student_name} (Риск: {risk_prob}%)"):
                        if interp['Приоритетные предметы']:
                            st.write("**Приоритетные предметы:**")
                            for item in interp['Приоритетные предметы']:
                                st.write(f"**{item['предмет']}**")
                                st.write(f"   - Оценка: {item['оценка']}")
                                st.write(f"   - Количество зависимых дисциплин: {item['количество_зависимых']}")
                                if item['зависимые_дисциплины']:
                                    deps_str = ", ".join(item['зависимые_дисциплины'])
                                    st.write(f"   - Дисциплины: {deps_str}")
                                st.write("")
                
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    results_df.to_excel(writer, sheet_name='Анализ рисков')
                
                st.download_button(
                    label="Скачать отчет в формате Excel",
                    data=buffer.getvalue(),
                    file_name=f"Отчет_рисков_семестр_{semester_choice}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("Студентов с вероятностью академического риска 50% и более не обнаружено.")
